"""
Generate Excel file with all 28 P6 prototype cards (v4: 敵人四類分類).
Two sheets:
1. 卡牌資料 - tabular reference (templates with pool_size + lock_class shown)
2. 卡片列印 - card-formatted print layout (28 physical cards, lock_class visible per card)

v4 changes (2026-05-11):
- 警示(intel_class): 從揭露精英化改為揭露敵人類別(兔子/狼/豹/熊)
- 試探擊(compound_generic): 從揭露 is_elite 改為揭露 is_failure_game_over
- 全知之眼(character_omniscient): 揭露範圍改為包含 enemy_class
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# 28 cards (templates):
# (id, name, resource, function, form, weakness, contribution_text, level,
#  pool_size, use_limit, lock_class, special, desc)
#
# lock_class:
#   'a' = none (Lock 與否不影響;不寫 lock 標示)
#   'b' = optional (情報卡;不 lock = 廣度,Lock = 深度)
#   'c' = required (強制 lock 才生效;flexible 卡 lock 時分配)
templates = [
    # Layer A: 衝擊池(a 類)
    ('tool_pool_impact_basic', '直擊', '工具', '戰鬥', '池', '衝擊',
     '+1 衝擊', '★', 3, 3, 'a', '', '對單一目標貢獻 +1 衝擊'),
    ('tool_pool_impact_strong', '重斬', '工具', '戰鬥', '池', '衝擊',
     '+2 衝擊', '★★', 2, 2, 'a', '', '對單一目標貢獻 +2 衝擊'),

    # Layer A: 穿刺池(a 類)
    ('tool_pool_pierce_basic', '銳刺', '工具', '戰鬥', '池', '穿刺',
     '+1 穿刺', '★', 3, 3, 'a', '', '對單一目標貢獻 +1 穿刺'),
    ('tool_pool_pierce_strong', '貫穿', '工具', '戰鬥', '池', '穿刺',
     '+2 穿刺', '★★', 2, 2, 'a', '', '對單一目標貢獻 +2 穿刺'),

    # Layer A: 燃燒池(a 類)
    ('tool_pool_burn_basic', '點火', '工具', '戰鬥', '池', '燃燒',
     '+1 燃燒', '★', 2, 3, 'a', '', '對單一目標貢獻 +1 燃燒'),
    ('tool_pool_burn_strong', '預判燒灼', '工具', '戰鬥', '池', '燃燒',
     '+2 燃燒', '★★', 2, 2, 'a', '', '對單一目標貢獻 +2 燃燒'),

    # Layer B: 工具 × 情報(b 類,雙模式)
    ('intel_weakness', '偵察', '工具', '情報', '個', '無',
     '揭露', '★', 1, 2, 'b',
     '❒ 不 lock: 結束本擊時揭露此連戰所有剩餘敵人的弱點類型\n❑ Lock: 即時揭露當前敵人弱點類型',
     '情報卡。不 lock 換廣度,Lock 換深度'),
    ('intel_requirements', '估量', '工具', '情報', '個', '無',
     '揭露', '★', 1, 2, 'b',
     '❒ 不 lock: 揭露此連戰所有剩餘敵人的最低需求次數\n❑ Lock: 揭露當前敵人完整需求表',
     '情報卡。不 lock 看剩餘最低,Lock 看當前完整'),
    ('intel_class', '警示', '工具', '情報', '個', '無',
     '揭露', '★', 1, 2, 'b',
     '❒ 不 lock: 揭露此連戰所有剩餘敵人的類別(兔子/狼/豹/熊)\n❑ Lock: 揭露當前敵人類別',
     '情報卡。揭露敵人類別,告訴你「失敗會怎樣」'),

    # Layer B: 工具 × 複合(試探)(c 類,binary 揭露)
    ('compound_impact', '試探斬', '工具', '複合', '個', '衝擊',
     '+1 衝擊', '★', 1, 3, 'c',
     '🔒 Lock 觸發: 揭露當前敵人弱點是否為衝擊(yes/no)',
     '試探卡(必 lock)。+1 衝擊 + 驗證弱點是否為衝擊'),
    ('compound_pierce', '探刺', '工具', '複合', '個', '穿刺',
     '+1 穿刺', '★', 1, 3, 'c',
     '🔒 Lock 觸發: 揭露當前敵人最低需求是否 ≤ 2(yes/no)',
     '試探卡(必 lock)。+1 穿刺 + 驗證敵人是否易擊殺'),
    ('compound_generic', '試探擊', '工具', '複合', '個', '通用',
     '+1 任一', '★', 1, 3, 'c',
     '🔒 Lock 觸發: 分配 +1 任一 + 揭露當前敵人失敗是否會 GAME OVER(yes/no)',
     '試探卡(必 lock + 分配)。+1 任一 + 驗證失敗是否致命'),

    # Layer B: 爆發(a 類 / c 類)
    ('burst_impact', '重擊', '爆發', '戰鬥', '個', '衝擊',
     '+3 衝擊', '★★★', 1, 1, 'a',
     '', '對單一目標貢獻 +3 衝擊'),
    ('burst_pierce', '致命刺', '爆發', '戰鬥', '個', '穿刺',
     '+3 穿刺', '★★★', 1, 1, 'a',
     '', '對單一目標貢獻 +3 穿刺'),
    ('burst_burn', '烈焰', '爆發', '戰鬥', '個', '燃燒',
     '+3 燃燒', '★★★', 1, 1, 'a',
     '', '對單一目標貢獻 +3 燃燒'),
    ('burst_flexible', '爆裂', '爆發', '戰鬥', '個', '通用',
     '+2 任一', '★★', 1, 1, 'c',
     '🔒 Lock 必須(分配): Lock 時必須選擇傷害類型分配 +2',
     '對單一目標貢獻 +2 任一類型(必 lock 分配)'),

    # Layer B: 角色(a 類 / c 類)
    ('character_dual', '殺戮意志', '角色', '戰鬥', '個', '通用',
     '+2 衝擊 +2 穿刺', '★★★', 1, 1, 'a',
     '', '戰役招牌:同時貢獻 +2 衝擊與 +2 穿刺'),
    ('character_focus', '致命凝視', '角色', '戰鬥', '個', '穿刺',
     '+4 穿刺', '★★★', 1, 1, 'a',
     '對連戰最高需求者所有貢獻 +1', '+4 穿刺;對最高需求者+1'),
    ('character_omniscient', '全知之眼', '角色', '情報', '個', '無',
     '揭露', '★★★', 1, 1, 'c',
     '🔒 Lock 必須: 揭露此連戰剩餘所有敵人的完整資訊(弱點+需求表+敵人類別+戰鬥狀態 instance)',
     '戰役招牌:揭露剩餘所有敵人完整資訊 + 戰鬥狀態(情報卡無法取得)'),
    ('character_insight', '洞察一擊', '角色', '複合', '個', '通用',
     '+2 任一', '★★★', 1, 1, 'c',
     '🔒 Lock 必須(分配): 分配 +2 任一 + 揭露當前敵人弱點',
     '戰役招牌:+2 任一(分配)+ Lock 揭露當前弱點'),
]

# Color schemes
RESOURCE_COLORS = {
    '工具': 'D9E1F2',  # light blue
    '爆發': 'FCE4D6',  # light orange
    '角色': 'E2EFDA',  # light green
}

WEAKNESS_COLORS = {
    '衝擊': 'FFC7CE',  # pink
    '穿刺': 'BDD7EE',  # blue
    '燃燒': 'FFD966',  # yellow
    '通用': 'D9D9D9',  # gray
    '無':   'F2F2F2',  # very light gray
}

FUNCTION_COLORS = {
    '戰鬥': 'E2EFDA',  # green
    '情報': 'D9E1F2',  # blue
    '複合': 'FFF2CC',  # yellow
}

FORM_COLORS = {
    '池': 'EDEDED',  # light gray (resource pool)
    '個': 'FFFFFF',  # white (individual)
}

LOCK_CLASS_COLORS = {
    'a': 'FFFFFF',  # white (no lock needed)
    'b': 'D5E8D4',  # light green (optional lock)
    'c': 'F8CECC',  # light red (required lock)
}

LOCK_CLASS_DISPLAY = {
    'a': '',          # no marker
    'b': '⚙ 可選 Lock',  # gear icon for choice
    'c': '🔒 強制 Lock',  # lock icon for required
}

wb = Workbook()

# ============================================================
# Sheet 1: 卡牌資料 (template-level reference)
# ============================================================
ws1 = wb.active
ws1.title = '卡牌資料'

headers = ['#', 'ID', '名稱', '資源', '功能', '形態', '弱點類型',
           '貢獻', '★', 'pool', 'use', 'Lock 類', '特殊效果', '簡述']
ws1.append(headers)

# Header styling
header_fill = PatternFill('solid', fgColor='1F4E78')
header_font = Font(bold=True, color='FFFFFF', size=11)
for col_num in range(1, len(headers) + 1):
    cell = ws1.cell(1, col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Template rows
for i, t in enumerate(templates, 1):
    row = i + 1
    ws1.cell(row, 1, i)
    for j, val in enumerate(t, 2):
        ws1.cell(row, j, val)

# Column widths
widths = [4, 30, 12, 6, 6, 4, 6, 16, 5, 5, 5, 8, 50, 36]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Row heights
ws1.row_dimensions[1].height = 25
for r in range(2, len(templates) + 2):
    ws1.row_dimensions[r].height = 50  # taller for special_text wrap

# Borders
thin = Side(border_style='thin', color='808080')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for r in range(1, len(templates) + 2):
    for c in range(1, len(headers) + 1):
        ws1.cell(r, c).border = border

# Color code by resource class (整列)
for i, t in enumerate(templates, 1):
    resource = t[2]
    if resource in RESOURCE_COLORS:
        for c in range(1, len(headers) + 1):
            ws1.cell(i + 1, c).fill = PatternFill('solid', fgColor=RESOURCE_COLORS[resource])

# Color code weakness column
for i, t in enumerate(templates, 1):
    weakness = t[6]
    if weakness in WEAKNESS_COLORS:
        ws1.cell(i + 1, 7).fill = PatternFill('solid', fgColor=WEAKNESS_COLORS[weakness])

# Color code lock_class column
for i, t in enumerate(templates, 1):
    lock_class = t[10]
    if lock_class in LOCK_CLASS_COLORS:
        ws1.cell(i + 1, 12).fill = PatternFill('solid', fgColor=LOCK_CLASS_COLORS[lock_class])
        ws1.cell(i + 1, 12).value = lock_class.upper()
        ws1.cell(i + 1, 12).font = Font(bold=True, size=12)

# Wrap text and align
for r in range(2, len(templates) + 2):
    for c in range(1, len(headers) + 1):
        align = Alignment(wrap_text=True, vertical='center',
                          horizontal='left' if c >= 13 else 'center')
        ws1.cell(r, c).alignment = align

# Freeze panes
ws1.freeze_panes = 'D2'

# ============================================================
# Sheet 2: 卡片列印 (physical cards, pool templates duplicated by pool_size)
# ============================================================
ws2 = wb.create_sheet('卡片列印')

# Build physical card list (expand pool templates by pool_size)
physical_cards = []
for t in templates:
    pool_size = t[8] if t[4] == '池' else 1
    for copy_idx in range(pool_size):
        physical_cards.append((t, copy_idx + 1, pool_size))

# Layout: each card occupies 9 rows × 4 cols (added 1 row for lock_class line)
ROWS_PER_CARD = 9
COLS_PER_CARD = 4
CARDS_PER_ROW = 3

thick = Side(border_style='medium', color='000000')
thin_border = Side(border_style='thin', color='B0B0B0')

for idx, (t, copy_idx, pool_size) in enumerate(physical_cards):
    (card_id, name, resource, function, form, weakness,
     contribution, level, _pool, use_limit, lock_class, special, desc) = t

    card_row = idx // CARDS_PER_ROW
    card_col = idx % CARDS_PER_ROW

    start_row = card_row * (ROWS_PER_CARD + 1) + 1
    start_col = card_col * (COLS_PER_CARD + 1) + 1
    end_col = start_col + COLS_PER_CARD - 1
    end_row = start_row + ROWS_PER_CARD - 1

    # Row 1: Name (3 cols merged) | Level (1 col)
    name_text = name
    if form == '池':
        name_text = f'{name}  ({copy_idx}/{pool_size})'
    name_cell = ws2.cell(start_row, start_col, name_text)
    name_cell.font = Font(bold=True, size=14)
    name_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws2.merge_cells(start_row=start_row, end_row=start_row,
                    start_column=start_col, end_column=start_col + 2)

    level_cell = ws2.cell(start_row, start_col + 3, level)
    level_cell.font = Font(bold=True, size=12, color='B45F06')
    level_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Row 2: Resource | Function | Weakness | Form
    res_cell = ws2.cell(start_row + 1, start_col, resource)
    res_cell.fill = PatternFill('solid', fgColor=RESOURCE_COLORS.get(resource, 'FFFFFF'))
    res_cell.alignment = Alignment(horizontal='center', vertical='center')
    res_cell.font = Font(size=10, bold=True)

    func_cell = ws2.cell(start_row + 1, start_col + 1, function)
    func_cell.fill = PatternFill('solid', fgColor=FUNCTION_COLORS.get(function, 'FFFFFF'))
    func_cell.alignment = Alignment(horizontal='center', vertical='center')
    func_cell.font = Font(size=10, bold=True)

    weak_cell = ws2.cell(start_row + 1, start_col + 2, weakness)
    weak_cell.fill = PatternFill('solid', fgColor=WEAKNESS_COLORS.get(weakness, 'FFFFFF'))
    weak_cell.alignment = Alignment(horizontal='center', vertical='center')
    weak_cell.font = Font(size=10, bold=True)

    form_cell = ws2.cell(start_row + 1, start_col + 3, form)
    form_cell.fill = PatternFill('solid', fgColor=FORM_COLORS.get(form, 'FFFFFF'))
    form_cell.alignment = Alignment(horizontal='center', vertical='center')
    form_cell.font = Font(size=10, bold=True)

    # Row 3: Lock class banner (4 cols merged)
    lock_text = LOCK_CLASS_DISPLAY.get(lock_class, '')
    lock_color = '4F4F4F'  # default dark gray
    if lock_class == 'b':
        lock_color = '2E7D32'  # dark green
    elif lock_class == 'c':
        lock_color = 'C62828'  # dark red
    lock_cell = ws2.cell(start_row + 2, start_col, lock_text)
    lock_cell.fill = PatternFill('solid', fgColor=LOCK_CLASS_COLORS.get(lock_class, 'FFFFFF'))
    lock_cell.font = Font(size=10, bold=True, color=lock_color)
    lock_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws2.merge_cells(start_row=start_row + 2, end_row=start_row + 2,
                    start_column=start_col, end_column=start_col + 3)

    # Row 4: Contribution (3 cols merged) | Use dots (1 col)
    contrib_cell = ws2.cell(start_row + 3, start_col, contribution)
    contrib_cell.font = Font(size=12, bold=True, color='1F4E78')
    contrib_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws2.merge_cells(start_row=start_row + 3, end_row=start_row + 3,
                    start_column=start_col, end_column=start_col + 2)

    use_dots = '●' * use_limit
    use_text = f'{use_dots}\n({use_limit} use)'
    use_cell = ws2.cell(start_row + 3, start_col + 3, use_text)
    use_cell.font = Font(size=9, bold=True)
    use_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Row 5-6: Special effect (4 cols merged, 2 rows)
    sp_cell = ws2.cell(start_row + 4, start_col, special)
    sp_cell.font = Font(size=9, italic=True, color='C00000' if special else '808080')
    sp_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws2.merge_cells(start_row=start_row + 4, end_row=start_row + 5,
                    start_column=start_col, end_column=start_col + 3)

    # Row 7-8: Description (4 cols merged, 2 rows)
    d_cell = ws2.cell(start_row + 6, start_col, desc)
    d_cell.font = Font(size=9)
    d_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws2.merge_cells(start_row=start_row + 6, end_row=start_row + 7,
                    start_column=start_col, end_column=start_col + 3)

    # Row 9: ID (4 cols merged, small font)
    id_text = f'#{idx+1} · {card_id}'
    id_cell = ws2.cell(start_row + 8, start_col, id_text)
    id_cell.font = Font(size=8, color='808080')
    id_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws2.merge_cells(start_row=start_row + 8, end_row=start_row + 8,
                    start_column=start_col, end_column=start_col + 3)

    # Apply border to entire card
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws2.cell(r, c).border = Border(
                left=thick if c == start_col else thin_border,
                right=thick if c == end_col else thin_border,
                top=thick if r == start_row else thin_border,
                bottom=thick if r == end_row else thin_border,
            )

# Set column widths for card layout
card_col_widths = [11, 11, 11, 11]
for card_col in range(CARDS_PER_ROW):
    base = card_col * (COLS_PER_CARD + 1) + 1
    for offset, width in enumerate(card_col_widths):
        ws2.column_dimensions[get_column_letter(base + offset)].width = width
    sep_col = base + COLS_PER_CARD
    if card_col < CARDS_PER_ROW - 1:
        ws2.column_dimensions[get_column_letter(sep_col)].width = 2

# Set row heights (9 rows per card now)
total_card_rows = (len(physical_cards) + CARDS_PER_ROW - 1) // CARDS_PER_ROW
for card_row in range(total_card_rows):
    base_row = card_row * (ROWS_PER_CARD + 1) + 1
    # Card rows: name, type, lock_class, contribution, special1, special2, desc1, desc2, id
    heights = [25, 22, 18, 26, 22, 22, 18, 18, 14]
    for offset, h in enumerate(heights):
        ws2.row_dimensions[base_row + offset].height = h
    sep_row = base_row + ROWS_PER_CARD
    if card_row < total_card_rows - 1:
        ws2.row_dimensions[sep_row].height = 8

# Page setup for printing
ws2.page_setup.paper_size = ws2.PAPERSIZE_A4
ws2.page_setup.orientation = ws2.ORIENTATION_PORTRAIT
ws2.page_margins.left = 0.4
ws2.page_margins.right = 0.4
ws2.page_margins.top = 0.4
ws2.page_margins.bottom = 0.4
ws2.print_options.horizontalCentered = True

# Page break: 9 cards per page (3x3)
for page_break_card_row in range(3, total_card_rows, 3):
    page_break_row = page_break_card_row * (ROWS_PER_CARD + 1)
    ws2.row_breaks.append(openpyxl.worksheet.pagebreak.Break(id=page_break_row))

# Save
output_path = r'f:\CCTEST\haloflag ai trpg\P6\docs\卡牌規格.xlsx'
wb.save(output_path)
print(f'Created: {output_path}')
print(f'Templates: {len(templates)}')
print(f'Physical cards (pool expanded): {len(physical_cards)}')
print(f'Sheets: {wb.sheetnames}')
print(f'Lock class distribution: a={sum(1 for t in templates if t[10]=="a")}, '
      f'b={sum(1 for t in templates if t[10]=="b")}, '
      f'c={sum(1 for t in templates if t[10]=="c")}')
